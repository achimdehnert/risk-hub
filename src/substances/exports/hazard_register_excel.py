# src/substances/exports/hazard_register_excel.py
"""
Gefahrstoffverzeichnis Export als Excel.

Erfüllt die Anforderungen nach GefStoffV §6.
"""

import io
from typing import Optional
from uuid import UUID

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_hazard_register_excel(
    tenant_id: UUID,
    site_id: Optional[UUID] = None
) -> io.BytesIO:
    """
    Generiert Gefahrstoffverzeichnis nach GefStoffV §6 als Excel.

    Args:
        tenant_id: Tenant-ID
        site_id: Optional - filtert nach Standort

    Returns:
        BytesIO Buffer mit Excel-Datei
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl ist nicht installiert. "
            "Bitte 'pip install openpyxl' ausführen."
        )

    from django.db.models import Q
    from substances.models import Substance

    wb = Workbook()
    ws = wb.active
    ws.title = "Gefahrstoffverzeichnis"

    # Styles
    header_fill = PatternFill(start_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    cmr_fill = PatternFill(start_color="FFCCCC", fill_type="solid")

    # Header
    headers = [
        "Nr.",
        "Stoffname",
        "CAS-Nr.",
        "Handelsname",
        "Hersteller",
        "Lagerklasse",
        "CMR",
        "Menge",
        "Einheit",
        "Lagerort",
        "H-Sätze",
        "P-Sätze",
        "Piktogramme",
        "SDS-Datum",
        "SDS-Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Daten abfragen
    query = Q(tenant_id=tenant_id, status="active")

    if site_id:
        substances = Substance.objects.filter(
            query,
            inventory_items__site_id=site_id
        ).distinct().select_related(
            "manufacturer"
        ).prefetch_related(
            "identifiers",
            "sds_revisions__hazard_statements",
            "sds_revisions__precautionary_statements",
            "sds_revisions__pictograms",
            "inventory_items",
        )
    else:
        substances = Substance.objects.filter(
            query
        ).select_related(
            "manufacturer"
        ).prefetch_related(
            "identifiers",
            "sds_revisions__hazard_statements",
            "sds_revisions__precautionary_statements",
            "sds_revisions__pictograms",
            "inventory_items",
        )

    # Zeilen schreiben
    for row_num, substance in enumerate(substances, 2):
        current_sds = substance.current_sds
        inventory = substance.inventory_items.first()

        # Nr.
        ws.cell(row=row_num, column=1, value=row_num - 1)

        # Stoffname
        ws.cell(row=row_num, column=2, value=substance.name)

        # CAS-Nr.
        ws.cell(row=row_num, column=3, value=substance.cas_number or "")

        # Handelsname
        ws.cell(row=row_num, column=4, value=substance.trade_name)

        # Hersteller
        manufacturer_name = ""
        if substance.manufacturer:
            manufacturer_name = substance.manufacturer.name
        ws.cell(row=row_num, column=5, value=manufacturer_name)

        # Lagerklasse
        ws.cell(row=row_num, column=6, value=substance.storage_class)

        # CMR
        cmr_cell = ws.cell(row=row_num, column=7)
        cmr_cell.value = "Ja" if substance.is_cmr else "Nein"
        if substance.is_cmr:
            cmr_cell.fill = cmr_fill

        # Menge & Einheit & Lagerort
        if inventory:
            ws.cell(row=row_num, column=8, value=float(inventory.quantity))
            ws.cell(row=row_num, column=9, value=inventory.unit)
            ws.cell(row=row_num, column=10, value=inventory.storage_location)
        else:
            ws.cell(row=row_num, column=8, value="")
            ws.cell(row=row_num, column=9, value="")
            ws.cell(row=row_num, column=10, value="")

        # H-Sätze, P-Sätze, Piktogramme, SDS-Datum, SDS-Status
        if current_sds:
            h_codes = ", ".join(
                h.code for h in current_sds.hazard_statements.all()
            )
            p_codes = ", ".join(
                p.code for p in current_sds.precautionary_statements.all()
            )
            pictogram_codes = ", ".join(
                p.code for p in current_sds.pictograms.all()
            )
            sds_date = current_sds.revision_date.strftime("%d.%m.%Y")
            sds_status = current_sds.get_status_display()
        else:
            h_codes = ""
            p_codes = ""
            pictogram_codes = ""
            sds_date = ""
            sds_status = "Kein SDS"

        ws.cell(row=row_num, column=11, value=h_codes)
        ws.cell(row=row_num, column=12, value=p_codes)
        ws.cell(row=row_num, column=13, value=pictogram_codes)
        ws.cell(row=row_num, column=14, value=sds_date)
        ws.cell(row=row_num, column=15, value=sds_status)

        # Border für alle Zellen
        for col in range(1, 16):
            ws.cell(row=row_num, column=col).border = thin_border

    # Spaltenbreiten anpassen
    column_widths = [
        5,   # Nr.
        30,  # Stoffname
        15,  # CAS-Nr.
        25,  # Handelsname
        20,  # Hersteller
        10,  # Lagerklasse
        6,   # CMR
        10,  # Menge
        8,   # Einheit
        20,  # Lagerort
        30,  # H-Sätze
        40,  # P-Sätze
        20,  # Piktogramme
        12,  # SDS-Datum
        12,  # SDS-Status
    ]

    for i, width in enumerate(column_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
        ws.column_dimensions[col_letter].width = width

    # Filter aktivieren
    ws.auto_filter.ref = f"A1:O{ws.max_row}"

    # Zeile 1 fixieren
    ws.freeze_panes = "A2"

    # In Buffer speichern
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
